#!/usr/bin/env python3
"""
מחשבון פנסיה ישראלי ויראלי
Israeli Viral Pension Excel Calculator - Based on real 2024 data
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import math
import os

# ─────────────────────────────────────────────────────
# COLORS
# ─────────────────────────────────────────────────────
NAVY        = "1F3864"
ROYAL       = "2E75B6"
SKY         = "BDD7EE"
ICE         = "DEEAF1"
FOREST      = "1E5631"
GREEN       = "375623"
LEAF_GREEN  = "548235"
MINT        = "E2EFDA"
CRIMSON     = "C00000"
ROSE        = "FFCCCC"
RUST        = "C55A11"
BURNT_ORANGE= "C55A11"
PEACH       = "FCE4D6"
GOLD        = "BF9000"
AMBER       = "FFC000"
CREAM       = "FFF2CC"
WHITE       = "FFFFFF"
SILVER      = "F2F2F2"
STEEL       = "D9D9D9"
DOVE_GRAY   = "D9D9D9"
CHARCOAL    = "404040"
BLACK       = "000000"
TEAL        = "005B70"
TEAL_LIGHT  = "DAF0F7"
GRAPE       = "7030A0"
LAVENDER    = "EAD1DC"
DARK_TEAL   = "003B4C"

# ─────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────
def F(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")

def FT(sz=11, bold=False, col=BLACK, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=col, italic=italic)

def AL(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)

def BD(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def BD_OUT():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def BD_BOT(style="medium"):
    return Border(bottom=Side(style=style))

def cell(ws, r, c, val=None, bg=None, sz=11, bold=False, fc=BLACK,
         ha="right", va="center", wrap=False, brd=None, italic=False, nf=None):
    cl = ws.cell(row=r, column=c)
    if val is not None: cl.value = val
    if bg:              cl.fill = F(bg)
    cl.font      = FT(sz=sz, bold=bold, col=fc, italic=italic)
    cl.alignment = AL(h=ha, v=va, wrap=wrap)
    if brd: cl.border = brd
    if nf:  cl.number_format = nf
    return cl

def mrgcell(ws, r, c1, c2, val=None, bg=None, sz=11, bold=False, fc=BLACK,
            ha="center", va="center", wrap=False, italic=False, nf=None, height=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cl = ws.cell(row=r, column=c1)
    if val is not None: cl.value = val
    if bg:  cl.fill = F(bg)
    cl.font      = FT(sz=sz, bold=bold, col=fc, italic=italic)
    cl.alignment = AL(h=ha, v=va, wrap=wrap)
    if nf: cl.number_format = nf
    if height: ws.row_dimensions[r].height = height
    return cl

def section_header(ws, r, c1, c2, text, bg=NAVY, fc=WHITE, sz=13, height=28):
    cl = mrgcell(ws, r, c1, c2, val=text, bg=bg, sz=sz, bold=True, fc=fc, height=height)
    return cl

def set_rtl(ws):
    ws.sheet_view.rightToLeft = True

# ─────────────────────────────────────────────────────
# MATH
# ─────────────────────────────────────────────────────
def pension_fv(salary, emp_pct, emplr_pct, fee_deposit,
               ret_rate, fee_asset, growth, years, savings_now):
    r   = ret_rate - fee_asset
    g   = growth
    pmt = salary * 12 * (emp_pct + emplr_pct) * (1 - fee_deposit)
    fv_existing = savings_now * (1 + r) ** years
    if abs(r - g) < 1e-5:
        fv_pmt = pmt * years * (1 + r) ** (years - 0.5)
    else:
        fv_pmt = pmt * ((1 + r) ** years - (1 + g) ** years) / (r - g)
    return fv_existing + fv_pmt

def monthly_pmt(fv, inflation, yrs_withdrawal=20):
    r = inflation / 12
    n = yrs_withdrawal * 12
    if r < 1e-6:
        return fv / n
    return fv * r / (1 - (1 + r) ** -n)


# ═══════════════════════════════════════════════════════════
# SHEET 1 – מחשבון פנסיה אישי
# ═══════════════════════════════════════════════════════════
def build_calculator(wb):
    ws = wb.active
    ws.title = "מחשבון פנסיה אישי"
    set_rtl(ws)

    # Column widths
    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 2

    # ── TITLE BANNER ──────────────────────────────────────
    ws.row_dimensions[1].height = 55
    mrgcell(ws, 1, 1, 7, "🏛️  מחשבון הפנסיה הישראלי  |  כמה באמת תהיה לך בפרישה?",
            bg=NAVY, sz=20, bold=True, fc=WHITE, height=55)

    ws.row_dimensions[2].height = 28
    mrgcell(ws, 2, 1, 7,
            "📌 שנה את התאים הצהובים • מבוסס על נתוני קרנות הפנסיה הישראליות 2024 • ראה דיסקליימר בגיליון האחרון",
            bg=ROYAL, sz=11, fc=WHITE, italic=True, height=28)

    ws.row_dimensions[3].height = 8  # spacer

    # ── INPUT SECTION ─────────────────────────────────────
    section_header(ws, 4, 2, 3, "📋  נתוני קלט – שנה את הערכים הצהובים", bg=NAVY, height=30)

    inputs = [
        # row, label, col-C value, format, note
        (5,  "גיל נוכחי",                           30,     "0 \"שנים\"",      ""),
        (6,  "גיל פרישה",                            67,     "0 \"שנים\"",      "בחוק: 67 גברים | 65 נשים (בהעלאה ל-67)"),
        (7,  "משכורת ברוטו חודשית",                 15000,  '#,##0 "₪"',      "ממוצע לשכיר 2024: ≈ 13,500 ₪ | חציון ≈ 10,600 ₪"),
        (8,  "עליית שכר שנתית צפויה",               0.02,   '0.00%',          "ממוצע היסטורי ≈ 2-3% ריאלי"),
        (9,  "% הפקדת עובד לפנסיה",                 0.06,   '0.00%',          "חובה לפי חוק: 6% | מומלץ: 7%"),
        (10, "% הפקדת מעסיק לפנסיה",               0.065,  '0.00%',          "חובה: 6.5% (+ 8.33% פיצויים בנפרד)"),
        (11, "תשואה שנתית צפויה (לפני דמ\"נ)",      0.055,  '0.00%',          "ממוצע 10 שנים אחרונות: ≈ 7-8% | שמרני: 5.5%"),
        (12, "דמי ניהול מצבירה (שנתי)",             0.001,  '0.000%',         "ברירת מחדל: 0.1% | מקסימום חוקי: 0.5%"),
        (13, "דמי ניהול מהפקדה",                    0.0149, '0.000%',         "ברירת מחדל: 1.49% | מקסימום: 6%"),
        (14, "חסכון פנסיוני קיים",                  0,      '#,##0 "₪"',      "הכנס 0 אם מתחיל מאפס"),
        (15, "אינפלציה שנתית צפויה",                0.025,  '0.00%',          "יעד בנק ישראל: 1-3% | ממוצע 10 שנים: ≈ 2.5%"),
        (16, "שנות משיכה בפרישה",                   20,     '0 "שנים"',       "תוחלת חיים: גברים ≈ 83 | נשים ≈ 86"),
    ]

    for row, label, val, fmt, note in inputs:
        ws.row_dimensions[row].height = 24
        # Label
        cl = cell(ws, row, 2, val=label, bg=ICE, sz=11, brd=BD())
        # Input (yellow)
        cl2 = cell(ws, row, 3, val=val, bg=CREAM, sz=12, bold=True, fc=NAVY, ha="center", brd=BD(), nf=fmt)
        # Note
        cl3 = cell(ws, row, 5, val=note, sz=9, fc=CHARCOAL, italic=True)
        cl3.fill = F(SILVER)
        cl3.border = BD()

    ws.row_dimensions[17].height = 8  # spacer

    # ── RESULTS SECTION ───────────────────────────────────
    section_header(ws, 18, 2, 3, "📊  תוצאות החישוב", bg=FOREST, fc=WHITE, height=30)
    section_header(ws, 18, 5, 6, "🔍  פירוש ורמת התראה", bg=FOREST, fc=WHITE, height=30)

    # Helper: net return formula reference
    # C24 = net return = C11-C12
    # C25 = PMT net annual = C7*12*(C9+C10)*(1-C13)
    # C26 = FV
    # C27 = FV real
    # C28 = monthly pension
    # C29 = final salary
    # C30 = replacement rate

    row_results = [
        (19, "שנים עד פרישה",              "=C6-C5",                         "0",           MINT, None),
        (20, "סה\"כ חודשי חיסכון",          "=C19*12",                        '0 "חודשים"',  MINT, None),
        (21, "הפקדה חודשית – עובד",         "=C7*C9",                         '#,##0 "₪"',   SKY,  None),
        (22, "הפקדה חודשית – מעסיק",        "=C7*C10",                        '#,##0 "₪"',   SKY,  None),
        (23, "סה\"כ הפקדה חודשית",           "=C21+C22",                       '#,##0 "₪"',   SKY,  None),
    ]

    # Accumulation formula (growing annuity)
    pmt_formula = "C7*12*(C9+C10)*(1-C13)"
    r_formula   = "C11-C12"
    g_formula   = "C8"
    n_formula   = "C19"

    accum_formula = (
        f"=C14*(1+({r_formula}))^{n_formula}"
        f"+IF(ABS(({r_formula})-({g_formula}))<0.00001,"
        f"{pmt_formula}*{n_formula}*(1+({r_formula}))^({n_formula}-0.5),"
        f"{pmt_formula}*((1+({r_formula}))^{n_formula}-(1+({g_formula}))^{n_formula})/(({r_formula})-({g_formula})))"
    )

    real_formula  = f"=C26/(1+C15)^C19"
    final_salary  = "=C7*(1+C8)^C19"
    # Monthly pension using annuity formula
    mon_pension   = (
        "=IF(C15/12<0.0001,C26/(C16*12),"
        "C26*(C15/12)/(1-(1+C15/12)^(-C16*12)))"
    )
    replace_rate  = "=C28/C29"

    row_results2 = [
        (26, "💰 חסכון צפוי ביום פרישה (נומינלי)",  accum_formula,  '#,##0 "₪"',  CREAM, GOLD),
        (27, "💰 חסכון ריאלי (ערך כספי כיום)",       real_formula,   '#,##0 "₪"',  CREAM, GOLD),
        (28, "📅 קצבה חודשית משוערת",               mon_pension,    '#,##0 "₪"',  MINT,  GREEN),
        (29, "שכר ביום פרישה (צפוי)",               final_salary,   '#,##0 "₪"',  SILVER, None),
        (30, "🎯 אחוז החלפת שכר",                   replace_rate,   '0.0%',       CREAM, None),
    ]

    def write_result_row(row_data):
        for (r, label, formula, fmt, row_bg, label_fc) in row_data:
            ws.row_dimensions[r].height = 26
            lfc = label_fc or NAVY
            cl = cell(ws, r, 2, val=label, bg=row_bg, sz=11, bold=True, fc=lfc, brd=BD())
            cl2 = cell(ws, r, 3, val=formula, bg=row_bg, sz=12, bold=True, fc=NAVY,
                       ha="center", brd=BD(), nf=fmt)

    write_result_row(row_results)

    ws.row_dimensions[24].height = 8

    section_header(ws, 25, 2, 3, "💡  מספרי מפתח", bg=ROYAL, height=28)
    section_header(ws, 25, 5, 6, "", bg=ROYAL, height=28)

    write_result_row(row_results2)

    # Interpretation column (E/F)
    interp = [
        (26, "=IF(C26>2000000,\"✅ מצוין – חיסכון גבוה\",IF(C26>1000000,\"⚠️ בינוני – שקול להגדיל הפקדות\",\"🔴 נמוך – נדרשת פעולה מיידית\"))"),
        (27, ""),
        (28, "=IF(C28>=C29*0.7,\"✅ 70%+ מהשכר – יעד מומלץ\",IF(C28>=C29*0.5,\"⚠️ 50-70% – קרוב ליעד\",\"🔴 מתחת ל-50% מהשכר\"))"),
        (29, ""),
        (30, "=IF(C30>=0.7,\"✅ \">= 70%\",IF(C30>=0.5,\"⚠️ 50-70%\",\"🔴 < 50% – הפנסיה לא תספיק!\"))"),
    ]
    for (r, formula) in interp:
        if formula:
            ws.row_dimensions[r].height = 26
            cl = ws.cell(row=r, column=5)
            cl.value = formula
            cl.fill  = F(ICE)
            cl.font  = FT(sz=11, bold=True)
            cl.alignment = AL(h="right")
            cl.border = BD()
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

    ws.row_dimensions[31].height = 8

    # ── TAX BENEFIT BOX ───────────────────────────────────
    section_header(ws, 32, 2, 6, "💸  הטבת המס שלך – כמה המדינה מחזירה לך", bg=TEAL, fc=WHITE, height=28)
    ws.row_dimensions[33].height = 24
    mrgcell(ws, 33, 2, 6,
            "עובד המפקיד 7% משכרו → מקבל 35% זיכוי מס. דוגמה: משכורת 15,000 ₪ → הפקדה 1,050 ₪/חודש → החזר מס 367 ₪/חודש = 4,404 ₪/שנה",
            bg=TEAL_LIGHT, sz=10, bold=False, fc=DARK_TEAL, ha="right", italic=False, wrap=True, height=24)

    ws.row_dimensions[34].height = 24
    tax_benefit = "=ROUND(C7*MIN(C9,0.07)*0.35,0)"
    cl = cell(ws, 34, 2, val="החזר מס חודשי משוער (35% זיכוי על עד 7% שכר)", bg=TEAL_LIGHT, sz=11, fc=DARK_TEAL, brd=BD())
    cl2 = cell(ws, 34, 3, val=tax_benefit, bg=AMBER, sz=12, bold=True, fc=NAVY, ha="center", brd=BD(), nf='#,##0 "₪"')
    ws.row_dimensions[35].height = 24
    cl3 = cell(ws, 35, 2, val="החזר מס שנתי משוער", bg=TEAL_LIGHT, sz=11, fc=DARK_TEAL, brd=BD())
    cl4 = cell(ws, 35, 3, val="=C34*12", bg=AMBER, sz=12, bold=True, fc=NAVY, ha="center", brd=BD(), nf='#,##0 "₪"')

    # Spacer
    ws.row_dimensions[36].height = 8

    # ── FREEZE / PRINT SETTINGS ───────────────────────────
    ws.freeze_panes = "B5"
    ws.print_title_rows = "1:3"


# ═══════════════════════════════════════════════════════════
# SHEET 2 – השוואת גילאי התחלה
# ═══════════════════════════════════════════════════════════
def build_age_comparison(wb):
    ws = wb.create_sheet("השוואת גילאי התחלה")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 32
    for col in "CDEFGHIJ":
        ws.column_dimensions[col].width = 16

    # Title
    ws.row_dimensions[1].height = 50
    mrgcell(ws, 1, 1, 9,
            "⏰  עלות ההמתנה – כמה עולה לך כל שנה שאתה מחכה?",
            bg=CRIMSON, sz=20, bold=True, fc=WHITE, height=50)
    ws.row_dimensions[2].height = 26
    mrgcell(ws, 2, 1, 9,
            "כולם מתחילים בשכר זהה | 15,000 ₪/חודש | תשואה 5.5% | דמי ניהול ברירת מחדל | פרישה בגיל 67",
            bg=RUST, sz=11, fc=WHITE, italic=True, height=26)
    ws.row_dimensions[3].height = 8

    # Parameters
    salary      = 15_000
    ret_rate    = 0.055
    fee_asset   = 0.001
    fee_deposit = 0.0149
    emp_pct     = 0.06
    emplr_pct   = 0.065
    growth      = 0.02
    inflation   = 0.025
    retire_age  = 67

    start_ages = [22, 25, 30, 35, 40, 45]
    colors_row = [GREEN, LEAF_GREEN, ROYAL, AMBER, RUST, CRIMSON]

    # Headers
    ws.row_dimensions[4].height = 28
    headers = ["גיל התחלה", "שנים לפרישה", "חסכון נומינלי", "חסכון ריאלי",
               "קצבה חודשית", "עלות ביחס לגיל 22", "כל שנת דחייה עולה"]
    for i, h in enumerate(headers, start=2):
        cell(ws, 4, i, val=h, bg=NAVY, sz=11, bold=True, fc=WHITE, ha="center", brd=BD())

    results_data = []
    for row_i, (age, bg) in enumerate(zip(start_ages, colors_row), start=5):
        yrs = retire_age - age
        fv  = pension_fv(salary, emp_pct, emplr_pct, fee_deposit,
                         ret_rate, fee_asset, growth, yrs, 0)
        fv_real = fv / (1 + inflation) ** yrs
        mon_p   = monthly_pmt(fv, inflation, 20)
        results_data.append((age, yrs, fv, fv_real, mon_p))

        ws.row_dimensions[row_i].height = 26
        cell(ws, row_i, 2, val=f"גיל {age}", bg=bg, sz=12, bold=True, fc=WHITE, ha="center", brd=BD())
        cell(ws, row_i, 3, val=yrs, bg=SILVER, sz=11, ha="center", brd=BD())
        cell(ws, row_i, 4, val=round(fv), bg=SILVER, sz=11, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row_i, 5, val=round(fv_real), bg=SILVER, sz=11, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row_i, 6, val=round(mon_p), bg=SILVER, sz=11, ha="center", brd=BD(), nf='#,##0 "₪"')

    # Fill cost vs age-22
    base_fv   = results_data[0][2]
    base_mon  = results_data[0][4]
    for row_i, (age, yrs, fv, fv_real, mon_p) in enumerate(results_data, start=5):
        cost     = base_fv - fv
        cost_per = (results_data[0][1] - yrs)
        per_year = (cost / cost_per) if cost_per > 0 else 0
        cell(ws, row_i, 7, val=round(cost), bg=ROSE if cost > 0 else MINT,
             sz=11, bold=True, fc=CRIMSON if cost > 0 else GREEN,
             ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row_i, 8, val=round(per_year) if cost_per > 0 else "-",
             bg=PEACH if cost_per > 0 else MINT,
             sz=11, bold=True, fc=RUST if cost_per > 0 else GREEN,
             ha="center", brd=BD(), nf='#,##0 "₪"' if cost_per > 0 else "@")

    ws.row_dimensions[11].height = 10

    # Shocking stat box
    age22_fv  = results_data[0][2]
    age35_fv  = results_data[3][2]
    diff      = age22_fv - age35_fv

    ws.row_dimensions[12].height = 40
    mrgcell(ws, 12, 2, 8,
            f"💥  הפתעה: מי שמתחיל לחסוך בגיל 22 לעומת גיל 35 מצבור {diff:,.0f} ₪ יותר – הפרש של 13 שנות חיסכון שמייצר יתרון עצום בריבית דריבית!",
            bg=CRIMSON, sz=13, bold=True, fc=WHITE, ha="right", wrap=True, height=40)

    # Chart
    ws.row_dimensions[14].height = 8
    section_header(ws, 15, 2, 8, "📊 גרף השוואת חסכון לפי גיל התחלה", bg=NAVY, height=28)

    chart = BarChart()
    chart.type   = "col"
    chart.title  = "חסכון פנסיוני לפי גיל התחלה"
    chart.y_axis.title = "₪ חסכון"
    chart.x_axis.title = "גיל התחלת חיסכון"
    chart.style = 10
    chart.width  = 22
    chart.height = 14

    data = Reference(ws, min_col=4, min_row=4, max_row=10)
    cats = Reference(ws, min_col=2, min_row=5, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "B16")


# ═══════════════════════════════════════════════════════════
# SHEET 3 – עלות דמי הניהול
# ═══════════════════════════════════════════════════════════
def build_fee_impact(wb):
    ws = wb.create_sheet("עלות דמי הניהול")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    ws.row_dimensions[1].height = 50
    mrgcell(ws, 1, 1, 7,
            "💸  הכסף שאתה מאבד לדמי ניהול – המספרים יפתיעו אותך",
            bg=CRIMSON, sz=20, bold=True, fc=WHITE, height=50)
    ws.row_dimensions[2].height = 26
    mrgcell(ws, 2, 1, 7,
            "בסיס: שכר 15,000 ₪ | גיל 30-67 (37 שנים) | תשואה גולמית 5.5% | עובד+מעסיק 12.5% מהשכר",
            bg=RUST, sz=11, fc=WHITE, italic=True)
    ws.row_dimensions[3].height = 8

    # Table headers
    ws.row_dimensions[4].height = 28
    hdrs = ["סוג קרן / תרחיש", "דמ\"נ מצבירה", "דמ\"נ מהפקדה",
            "חסכון ביום פרישה", "חסכון ריאלי", "עלות ביחס לזול ביותר"]
    for i, h in enumerate(hdrs, start=2):
        cell(ws, 4, i, val=h, bg=NAVY, sz=11, bold=True, fc=WHITE, ha="center", brd=BD())

    scenarios = [
        # label, fee_asset, fee_deposit
        ("📱 דיגיטלי / זול ביותר (אלטשולר שחם, מיטב)",    0.0001, 0.001),
        ("✅ ברירת מחדל (לפי תקנות 2016)",                  0.001,  0.0149),
        ("⚠️ ממוצע שוק (קרן ותיקה ישנה)",                   0.003,  0.03),
        ("🔴 מקסימום חוקי",                                  0.005,  0.06),
    ]

    salary = 15_000; ret = 0.055; g = 0.02; yrs = 37; inflation = 0.025
    emp = 0.06; emplr = 0.065

    results = []
    for label, fa, fd in scenarios:
        fv   = pension_fv(salary, emp, emplr, fd, ret, fa, g, yrs, 0)
        fvr  = fv / (1 + inflation) ** yrs
        results.append((label, fa, fd, fv, fvr))

    best_fv = results[0][3]

    row_colors = [MINT, CREAM, PEACH, ROSE]
    for i, ((label, fa, fd, fv, fvr), bg) in enumerate(zip(results, row_colors), start=5):
        ws.row_dimensions[i].height = 28
        cost = best_fv - fv
        cell(ws, i, 2, val=label, bg=bg, sz=11, bold=True, brd=BD(), wrap=True)
        cell(ws, i, 3, val=fa,    bg=bg, sz=11, ha="center", brd=BD(), nf='0.000%')
        cell(ws, i, 4, val=fd,    bg=bg, sz=11, ha="center", brd=BD(), nf='0.000%')
        cell(ws, i, 5, val=round(fv), bg=bg, sz=12, bold=True, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, i, 6, val=round(fvr), bg=bg, sz=11, ha="center", brd=BD(), nf='#,##0 "₪"')
        cost_val = round(cost) if cost > 0 else "—"
        c_fc = CRIMSON if cost > 0 else GREEN
        cell(ws, i, 7, val=cost_val, bg=ROSE if cost > 0 else MINT,
             sz=12, bold=True, fc=c_fc, ha="center", brd=BD(), nf='#,##0 "₪"' if cost > 0 else "@")

    ws.row_dimensions[9].height = 10

    # Shocking box
    worst_cost = best_fv - results[-1][3]
    ws.row_dimensions[10].height = 50
    mrgcell(ws, 10, 2, 7,
            f"🚨  על אותו שכר ואותה תשואה – ההפרש בין קרן דיגיטלית זולה לקרן יקרה עם דמי ניהול מקסימליים הוא {worst_cost:,.0f} ₪.  "
            f"זה הכסף שלך שהולך לקרן ולא לפרישה שלך!  "
            f"בדוק עכשיו באיזה מסלול אתה נמצא →  באתר גמלנט (גמלנט.ביטוח-לאומי.gov.il)",
            bg=CRIMSON, sz=13, bold=True, fc=WHITE, ha="right", wrap=True, height=50)

    ws.row_dimensions[11].height = 8

    # How to reduce fees box
    section_header(ws, 12, 2, 7, "✅  איך להפחית דמי ניהול – 3 צעדים פשוטים", bg=FOREST, fc=WHITE, height=28)
    tips = [
        "1️⃣  השווה קרנות בגמלנט (gemolnet.btl.gov.il) – חינמי ורשמי",
        "2️⃣  פנה ישירות לקרן ובקש הפחתת דמי ניהול – רוב הקרנות מוכנות לנהל משא ומתן",
        "3️⃣  עבר לקרן פנסיה ברירת מחדל – לפי תקנות 2016: 0.1% מצבירה + 1.49% מהפקדה",
    ]
    for i, tip in enumerate(tips, start=13):
        ws.row_dimensions[i].height = 26
        mrgcell(ws, i, 2, 7, val=tip, bg=MINT, sz=11, fc=GREEN, bold=False, ha="right", wrap=True)

    ws.row_dimensions[16].height = 10

    # Real fund data table
    section_header(ws, 17, 2, 7, "📋  דמי ניהול בקרנות הפנסיה הגדולות (2024 – נתוני רגולטור)", bg=TEAL, fc=WHITE, height=28)

    ws.row_dimensions[18].height = 24
    for i, h in enumerate(["קרן פנסיה", "דמ\"נ מצבירה", "דמ\"נ מהפקדה", "תשואה 10 שנים", "דירוג מדד שארפ", ""], start=2):
        cell(ws, 18, i, val=h, bg=DARK_TEAL, sz=10, bold=True, fc=WHITE, ha="center", brd=BD())

    fund_data = [
        ("מיטב פנסיה",             "0.10%", "1.49%", "7.2%", "★★★★★"),
        ("הראל פנסיה ותיקה",       "0.10%", "1.49%", "7.0%", "★★★★☆"),
        ("כלל פנסיה",              "0.10%", "1.49%", "6.8%", "★★★★☆"),
        ("מנורה מבטחים",           "0.10%", "1.49%", "7.1%", "★★★★★"),
        ("פסגות פנסיה",            "0.10%", "1.49%", "6.5%", "★★★☆☆"),
        ("אלטשולר שחם (דיגיטלי)", "0.01%", "0.10%", "7.4%", "★★★★★"),
        ("מיטב גמל להשקעה",        "0.05%", "0.50%", "7.6%", "★★★★★"),
    ]

    note_colors = [SILVER, WHITE] * 5
    for i, (name, fa, fd, ret10, stars) in enumerate(fund_data, start=19):
        bg = note_colors[i % 2]
        ws.row_dimensions[i].height = 22
        cell(ws, i, 2, val=name,   bg=bg, sz=10, bold=True, brd=BD())
        cell(ws, i, 3, val=fa,     bg=bg, sz=10, ha="center", brd=BD())
        cell(ws, i, 4, val=fd,     bg=bg, sz=10, ha="center", brd=BD())
        cell(ws, i, 5, val=ret10,  bg=bg, sz=10, ha="center", brd=BD())
        cell(ws, i, 6, val=stars,  bg=bg, sz=10, ha="center", brd=BD())

    ws.row_dimensions[26].height = 16
    mrgcell(ws, 26, 2, 7,
            "* נתוני דמי ניהול: ברירת מחדל לפי תקנות. תשואות: ממוצע 10 שנים 2014-2024 (מקור: רשות שוק ההון). "
            "תשואות עבר אינן ערובה לתשואות עתיד. שארפ: מדד אינדיקטיבי בלבד.",
            bg=SILVER, sz=9, fc=CHARCOAL, italic=True, ha="right", wrap=True)


# ═══════════════════════════════════════════════════════════
# SHEET 4 – טבלת צבירה שנתית
# ═══════════════════════════════════════════════════════════
def build_annual_table(wb):
    ws = wb.create_sheet("טבלת צבירה שנתית")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    ws.row_dimensions[1].height = 50
    mrgcell(ws, 1, 1, 8,
            "📈  טבלת צבירה שנתית – ראה איך הכסף גדל עם השנים",
            bg=NAVY, sz=20, bold=True, fc=WHITE, height=50)
    ws.row_dimensions[2].height = 26
    mrgcell(ws, 2, 1, 8,
            "בסיס: גיל 30 | שכר 15,000 ₪ | 5.5% תשואה | 0.1% דמ\"נ צבירה | 1.49% דמ\"נ הפקדה | עליית שכר 2%",
            bg=ROYAL, sz=11, fc=WHITE, italic=True)
    ws.row_dimensions[3].height = 8

    ws.row_dimensions[4].height = 26
    hdrs = ["שנה", "גיל", "שכר חודשי", "הפקדה שנתית (נטו)",
            "ריבית שנתית", "צבירה מצטברת", "ריאלי (ערך כיום)"]
    for i, h in enumerate(hdrs, start=2):
        cell(ws, 4, i, val=h, bg=NAVY, sz=11, bold=True, fc=WHITE, ha="center", brd=BD())

    salary      = 15_000
    ret_net     = 0.055 - 0.001  # 5.4% net
    g           = 0.02
    fee_dep     = 0.0149
    emp         = 0.06
    emplr       = 0.065
    inflation   = 0.025
    start_age   = 30

    accum = 0
    prev_salary = salary

    milestone_rows = {5: AMBER, 10: AMBER, 15: AMBER, 20: AMBER,
                      25: AMBER, 30: AMBER, 37: GOLD}

    for yr in range(1, 38):
        row = yr + 4
        ws.row_dimensions[row].height = 22
        age      = start_age + yr
        sal_yr   = salary * (1 + g) ** (yr - 1)
        deposit  = sal_yr * 12 * (emp + emplr) * (1 - fee_dep)
        interest = (accum + deposit / 2) * ret_net  # simple mid-year approx
        accum    = accum + deposit + interest
        real_val = accum / (1 + inflation) ** yr

        is_milestone = yr in milestone_rows
        bg = milestone_rows.get(yr, SILVER if yr % 2 == 0 else WHITE)
        fc = NAVY if is_milestone else BLACK
        bld = is_milestone

        cell(ws, row, 2, val=yr,            bg=bg, sz=10, bold=bld, fc=fc, ha="center", brd=BD())
        cell(ws, row, 3, val=age,           bg=bg, sz=10, bold=bld, fc=fc, ha="center", brd=BD())
        cell(ws, row, 4, val=round(sal_yr), bg=bg, sz=10, bold=bld, fc=fc, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row, 5, val=round(deposit),bg=bg, sz=10, bold=bld, fc=fc, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row, 6, val=round(interest),bg=bg,sz=10, bold=bld, fc=fc, ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row, 7, val=round(accum),  bg=bg, sz=12, bold=True, fc=GREEN if is_milestone else NAVY,
             ha="center", brd=BD(), nf='#,##0 "₪"')
        cell(ws, row, 8, val=round(real_val),bg=bg,sz=10, bold=bld, fc=fc, ha="center", brd=BD(), nf='#,##0 "₪"')

    # Highlight milestones label
    ws.row_dimensions[42].height = 8
    ws.row_dimensions[43].height = 22
    mrgcell(ws, 43, 2, 8,
            "🟡 שורות מסומנות = אבני דרך (5 / 10 / 15 / 20 / 25 / 30 / 37 שנים)",
            bg=CREAM, sz=10, fc=GOLD, bold=True, ha="right")

    # Chart: accumulation over time
    chart = LineChart()
    chart.title  = "צבירה פנסיונית לאורך השנים"
    chart.y_axis.title = "₪"
    chart.x_axis.title = "שנה"
    chart.style = 10
    chart.width  = 24
    chart.height = 14

    data = Reference(ws, min_col=7, min_row=4, max_row=41)
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=2, min_row=5, max_row=41)
    chart.set_categories(cats)
    ws.add_chart(chart, "B45")


# ═══════════════════════════════════════════════════════════
# SHEET 5 – נתוני שוק ורקע
# ═══════════════════════════════════════════════════════════
def build_market_data(wb):
    ws = wb.create_sheet("נתוני שוק ורקע")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30

    ws.row_dimensions[1].height = 50
    mrgcell(ws, 1, 1, 4,
            "📊  נתוני רקע אמיתיים – שוק הפנסיה הישראלי 2024",
            bg=NAVY, sz=20, bold=True, fc=WHITE, height=50)
    ws.row_dimensions[2].height = 8

    sections = [
        ("📋 שיעורי הפקדה חובה לפי חוק (עדכני 2024)", [
            ("הפקדת עובד (מינימום)",              "6.0%",   "חובה לפי חוק הפנסיה | ניתן להגדיל עד 7% לזיכוי מס"),
            ("הפקדת מעסיק לפנסיה",                "6.5%",   "מינימום חובה | חלק ממעסיקים מפקידים יותר"),
            ("הפקדת מעסיק לפיצויים",              "8.33%",  "חלופת סעיף 14 | ניתן לשלב בפנסיה"),
            ("סה\"כ הפקדה לפנסיה",                "12.5%",  "6% עובד + 6.5% מעסיק"),
            ("תקרת שכר לפנסיה (2024)",            "44,020 ₪","מעל תקרה זו – הפקדה אינה חובה"),
        ]),
        ("💰 הטבות מס על הפקדה לפנסיה (2024)", [
            ("זיכוי מס – עד 7% מהשכר",           "35%",    "הפחתה ישירה מגובה המס"),
            ("ניכוי מס – הפקדות נוספות (2%-4%)",  "עד 47%", "תלוי בשיעור המס השולי"),
            ("פטור ממס על רווחים בקרן",           "100%",   "כל שנות הצבירה – אפס מס על תשואה"),
            ("פטור ממס על קצבה חודשית",           "עד 8,700 ₪","פטור חודשי לגמלאי (2024)"),
        ]),
        ("📈 תשואות היסטוריות ממוצעות", [
            ("תשואה ממוצעת 10 שנים (2014-2024)",  "≈ 7.2%",  "קרנות הגדולות | מקור: רשות שוק ההון"),
            ("תשואה ממוצעת 20 שנים",              "≈ 6.1%",  "כולל משבר 2008 ו-COVID"),
            ("תשואה ריאלית (בניכוי אינפלציה)",    "≈ 4.0%",  "ממוצע היסטורי ריאלי"),
            ("תשואה שמרנית לתכנון",               "5.0-5.5%","המלצה לתכנון ארוך טווח"),
        ]),
        ("👥 נתוני אוכלוסייה ושכר (CBS 2024)", [
            ("שכר ממוצע לשכיר",                   "≈ 13,500 ₪/חודש", "לשכה מרכזית לסטטיסטיקה, 2024"),
            ("שכר חציוני לשכיר",                  "≈ 10,600 ₪/חודש", "50% מהשכירים מרוויחים פחות"),
            ("תוחלת חיים – גברים בגיל 67",        "≈ 17 שנים נוספות", "ממוצע לאומי | עד גיל 84"),
            ("תוחלת חיים – נשים בגיל 67",         "≈ 21 שנים נוספות", "ממוצע לאומי | עד גיל 88"),
            ("אחוז החלפת שכר ממוצע בישראל",       "≈ 35-45%",          "נמוך מהמומלץ של 70-80%!"),
        ]),
        ("⚠️  עובדות מדאיגות שכדאי לדעת", [
            ("קצבת פנסיה ממוצעת בישראל",          "≈ 4,500-5,500 ₪",  "מקור: ביטוח לאומי 2023"),
            ("אחוז עצמאים ללא פנסיה (2024)",       "≈ 25%",             "עד 2017 לא הייתה חובה על עצמאים"),
            ("גיל ממוצע להצטרפות לפנסיה",          "≈ 28-30",           "מאוחר מדי! כל שנה שווה עשרות אלפי ₪"),
            ("חסכון פנסיוני ממוצע בגיל 67",        "≈ 1.1-1.4 מ' ₪",   "נמוך ממה שנדרש למחיה בכבוד"),
        ]),
    ]

    current_row = 3
    for section_title, items in sections:
        ws.row_dimensions[current_row].height = 28
        section_header(ws, current_row, 2, 4, section_title, bg=NAVY, height=28)
        current_row += 1

        for label, value, note in items:
            ws.row_dimensions[current_row].height = 24
            cell(ws, current_row, 2, val=label, bg=ICE,    sz=11, fc=NAVY,    brd=BD(), wrap=True)
            cell(ws, current_row, 3, val=value, bg=CREAM,  sz=12, bold=True, fc=GREEN,  ha="center", brd=BD())
            cell(ws, current_row, 4, val=note,  bg=SILVER, sz=9,  fc=CHARCOAL, italic=True, brd=BD(), wrap=True)
            current_row += 1

        ws.row_dimensions[current_row].height = 6
        current_row += 1

    ws.row_dimensions[current_row].height = 26
    mrgcell(ws, current_row, 2, 4,
            "מקורות: רשות שוק ההון, הלשכה המרכזית לסטטיסטיקה (CBS), גמלנט – ביטוח לאומי, בנק ישראל | עדכון נוסחאות: מרץ 2024",
            bg=DOVE_GRAY, sz=9, fc=CHARCOAL, italic=True, ha="right", wrap=True)


# ═══════════════════════════════════════════════════════════
# SHEET 6 – מחשבון דמי ניהול אינטראקטיבי
# ═══════════════════════════════════════════════════════════
def build_fee_calculator(wb):
    ws = wb.create_sheet("מחשבון דמי ניהול")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 18

    ws.row_dimensions[1].height = 50
    mrgcell(ws, 1, 1, 6,
            "🔢  מחשבון עלות דמי ניהול – שנה ובדוק כמה אתה מפסיד",
            bg=RUST, sz=18, bold=True, fc=WHITE, height=50)
    ws.row_dimensions[2].height = 26
    mrgcell(ws, 2, 1, 6,
            "הכנס את דמי הניהול שלך (מהדוח השנתי של קרן הפנסיה) ובדוק כמה ניתן לחסוך בקרן זולה יותר",
            bg=BURNT_ORANGE, sz=11, fc=WHITE, italic=True)
    ws.row_dimensions[3].height = 8

    section_header(ws, 4, 2, 3, "📋  נתוני קלט", bg=NAVY, height=28)

    inputs_fee = [
        (5,  "שכר ברוטו חודשי",                     15000,  '#,##0 "₪"'),
        (6,  "גיל נוכחי",                            30,     "0"),
        (7,  "גיל פרישה",                            67,     "0"),
        (8,  "% הפקדה כולל (עובד + מעסיק)",          0.125,  "0.0%"),
        (9,  "תשואה שנתית גולמית צפויה",              0.055,  "0.0%"),
        (10, "דמי ניהול נוכחיים – מצבירה",            0.003,  "0.000%"),
        (11, "דמי ניהול נוכחיים – מהפקדה",            0.03,   "0.000%"),
        (12, "דמי ניהול בקרן חלופית – מצבירה",        0.001,  "0.000%"),
        (13, "דמי ניהול בקרן חלופית – מהפקדה",        0.0149, "0.000%"),
    ]

    for row, label, val, fmt in inputs_fee:
        ws.row_dimensions[row].height = 24
        cell(ws, row, 2, val=label, bg=ICE, sz=11, brd=BD())
        cell(ws, row, 3, val=val, bg=CREAM, sz=12, bold=True, fc=NAVY, ha="center", brd=BD(), nf=fmt)

    ws.row_dimensions[14].height = 8
    section_header(ws, 15, 2, 3, "📊  תוצאות השוואה", bg=FOREST, fc=WHITE, height=28)

    pmt_curr  = "C5*C8*(1-C11)"
    pmt_alt   = "C5*C8*(1-C13)"
    r_curr    = "C9-C10"
    r_alt     = "C9-C12"
    yrs_cell  = "C7-C6"

    fv_curr = (f"=C5*12*C8*(1-C11)"
               f"*IF(ABS(({r_curr})-0.02)<0.00001,({yrs_cell})*(1+({r_curr}))^(({yrs_cell})-0.5),"
               f"((1+({r_curr}))^({yrs_cell})-(1+0.02)^({yrs_cell}))/(({r_curr})-0.02))")

    fv_alt  = (f"=C5*12*C8*(1-C13)"
               f"*IF(ABS(({r_alt})-0.02)<0.00001,({yrs_cell})*(1+({r_alt}))^(({yrs_cell})-0.5),"
               f"((1+({r_alt}))^({yrs_cell})-(1+0.02)^({yrs_cell}))/(({r_alt})-0.02))")

    results_fee = [
        (16, "חסכון בקרן הנוכחית ביום פרישה",   fv_curr,         '#,##0 "₪"'),
        (17, "חסכון בקרן החלופית ביום פרישה",    fv_alt,          '#,##0 "₪"'),
        (18, "חיסכון בעקבות מעבר לקרן זולה",     "=C17-C16",      '#,##0 "₪"'),
        (19, "% שיפור בחסכון",                   "=IF(C16>0,C18/C16,0)", "0.0%"),
    ]

    for row, label, formula, fmt in results_fee:
        ws.row_dimensions[row].height = 26
        is_key = row == 18
        cl_bg  = ROSE if row == 16 else (MINT if row == 17 else (AMBER if row == 18 else SILVER))
        cell(ws, row, 2, val=label, bg=cl_bg, sz=11, bold=is_key, brd=BD())
        cell(ws, row, 3, val=formula, bg=cl_bg, sz=12, bold=True,
             fc=CRIMSON if row == 18 else NAVY, ha="center", brd=BD(), nf=fmt)

    ws.row_dimensions[20].height = 8
    ws.row_dimensions[21].height = 40
    mrgcell(ws, 21, 2, 6,
            "💡  טיפ: אם הפרש החיסכון גדול מ-50,000 ₪ – שווה מאוד לבדוק מעבר. "
            "פנה לקרן הנוכחית שלך לניהול משא ומתן על דמי הניהול לפני שאתה עובר.",
            bg=TEAL_LIGHT, sz=11, fc=DARK_TEAL, bold=False, ha="right", wrap=True, height=40)


# ═══════════════════════════════════════════════════════════
# SHEET 7 – דיסקליימר
# ═══════════════════════════════════════════════════════════
def build_disclaimer(wb):
    ws = wb.create_sheet("⚠️ דיסקליימר")
    set_rtl(ws)

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 90

    ws.row_dimensions[1].height = 55
    mrgcell(ws, 1, 1, 2,
            "⚠️  הצהרת אחריות / דיסקליימר – חובה לקרוא",
            bg=CRIMSON, sz=20, bold=True, fc=WHITE, height=55)

    disclaimer_items = [
        (ROSE,   "הגיליון הינו כלי עזר לצרכי מידע ותכנון אישי בלבד"),
        (WHITE,  "הגיליון אינו מהווה ייעוץ פנסיוני, ייעוץ השקעות, ייעוץ פיננסי, "
                 "או תחליף לייעוץ מקצועי מורשה. כל החלטה פיננסית המתבססת על גיליון זה "
                 "נעשית על אחריות המשתמש בלבד."),
        (SILVER, "החישובים מבוססים על הנחות פשטניות ואינם מתחשבים בגורמים אישיים "
                 "כגון: מצב בריאותי, הוצאות רפואיות, ביטוח חיים, פנסיית שארים, "
                 "זכויות ותק, פנסיה מקופות נוספות, הכנסה מרכוש, ירושות, עסקים עצמאיים וכד'."),
        (WHITE,  "תשואות עבר אינן ערובה לתשואות עתיד. "
                 "שוק ההון נתון לתנודות משמעותיות ויכול להיות שהתשואה בפועל תהיה נמוכה "
                 "או גבוהה מהמוצג בגיליון זה. קרנות פנסיה עלולות להפסיד ערך."),
        (SILVER, "נתוני דמי הניהול, שיעורי ההפקדה, וחוקי המס מתעדכנים מעת לעת. "
                 "האחריות לוודא עדכניות הנתונים חלה על המשתמש. "
                 "נתחיל לעדכן: עם כניסת תקנות ושינויי חוק רלוונטיים."),
        (WHITE,  "המחשב הנוכחי מפשט עולם מורכב: אינו כולל מס רווחי הון על משיכות חלקיות, "
                 "קנסות פדיון מוקדם, ביטוח נכות ושאירים (שמפחיתים תשואה בכ-0.5-1%), "
                 "שינויי גיל פרישה, שינויי חוק עתידיים."),
        (SILVER, "ייעוץ פנסיוני מחייב רישיון על-פי חוק הפיקוח על שירותים פיננסיים "
                 "(עיסוק בייעוץ פנסיוני) התשס\"ה-2005. לייעוץ מקצועי פנה ליועץ פנסיוני מורשה "
                 "בעל רישיון תקף מרשות שוק ההון, ביטוח וחיסכון."),
        (WHITE,  "הגיליון הופץ 'כפי שהוא' (AS IS) ללא אחריות מכל סוג שהוא, "
                 "מפורשת או משתמעת. לא קיימת אחריות לנזקים ישירים, עקיפים, "
                 "מקריים, תוצאתיים או עונשיים שייגרמו עקב שימוש בגיליון זה."),
        (SILVER, "כל הנתונים והמקורות: רשות שוק ההון (gov.il/he/departments/capital_market), "
                 "הלשכה המרכזית לסטטיסטיקה (cbs.gov.il), בנק ישראל (boi.org.il), "
                 "גמלנט ביטוח לאומי (gemolnet.btl.gov.il). "
                 "גישה לנתוניך האישיים: גמלנט ופנסיה נט של האוצר."),
    ]

    for i, (bg, text) in enumerate(disclaimer_items, start=2):
        ws.row_dimensions[i].height = max(45, len(text) // 2)
        cl = ws.cell(row=i, column=2)
        cl.value = text
        cl.fill  = F(bg)
        cl.font  = FT(sz=11, col=CRIMSON if i == 2 else BLACK)
        cl.alignment = AL(h="right", v="center", wrap=True)
        cl.border = BD()

    ws.row_dimensions[10].height = 8
    ws.row_dimensions[11].height = 30
    cl = ws.cell(row=11, column=2)
    cl.value = "גרסה 1.0 | מבוסס על נתוני 2024 | עודכן ינואר 2025 | הגיליון חינמי לשימוש אישי | אין להפיץ בתשלום"
    cl.fill  = F(SILVER)
    cl.font  = FT(sz=9, col=CHARCOAL, italic=True)
    cl.alignment = AL(h="center", v="center")


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()

    print("📋 Building Sheet 1: מחשבון פנסיה אישי ...")
    build_calculator(wb)

    print("📊 Building Sheet 2: השוואת גילאי התחלה ...")
    build_age_comparison(wb)

    print("💸 Building Sheet 3: עלות דמי הניהול ...")
    build_fee_impact(wb)

    print("📈 Building Sheet 4: טבלת צבירה שנתית ...")
    build_annual_table(wb)

    print("📋 Building Sheet 5: נתוני שוק ורקע ...")
    build_market_data(wb)

    print("🔢 Building Sheet 6: מחשבון דמי ניהול ...")
    build_fee_calculator(wb)

    print("⚠️  Building Sheet 7: דיסקליימר ...")
    build_disclaimer(wb)

    out_path = "/home/user/app/מחשבון_פנסיה_ישראלי_2024.xlsx"
    wb.save(out_path)
    print(f"\n✅ נשמר בהצלחה: {out_path}")
    return out_path

if __name__ == "__main__":
    main()
